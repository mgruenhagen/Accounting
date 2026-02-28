"""
openpyxl named styles and formatting constants for the cash flow report.

Color palette: professional dark blue / light blue / white scheme
suitable for financial reporting.
"""

from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    NamedStyle,
    PatternFill,
    Side,
)

# ── Colors ────────────────────────────────────────────────────────────────────
COLOR_DARK_BLUE = "1F4E79"      # Header rows
COLOR_MID_BLUE = "2E75B6"       # Section headers
COLOR_LIGHT_BLUE = "BDD7EE"     # Subtotal rows
COLOR_VERY_LIGHT_BLUE = "DEEAF1"  # Alternating detail rows
COLOR_WHITE = "FFFFFF"
COLOR_RED = "C00000"            # Reconciliation error
COLOR_GREEN = "375623"          # Reconciliation OK

# ── Number formats ────────────────────────────────────────────────────────────
# Accounting format: parentheses for negatives, dash for zero
FMT_ACCOUNTING = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'
FMT_ACCOUNTING_PRECISE = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
FMT_INTEGER = '#,##0_);(#,##0)'
FMT_DATE = 'MM/DD/YYYY'

# ── Borders ───────────────────────────────────────────────────────────────────
THIN = Side(style="thin", color="000000")
MEDIUM = Side(style="medium", color="000000")
DOUBLE = Side(style="double", color="000000")

BORDER_TOP_BOTTOM = Border(top=THIN, bottom=THIN)
BORDER_TOP = Border(top=THIN)
BORDER_BOTTOM_THIN = Border(bottom=THIN)
BORDER_BOTTOM_DOUBLE = Border(bottom=DOUBLE)
BORDER_TOP_DOUBLE_BOTTOM_DOUBLE = Border(top=DOUBLE, bottom=DOUBLE)


def register_styles(wb) -> None:
    """
    Register all named styles on the workbook.
    Must be called before any sheets are written.
    """
    _add(wb, _title_style())
    _add(wb, _report_header_style())
    _add(wb, _section_header_style())
    _add(wb, _body_style())
    _add(wb, _body_indent_style())
    _add(wb, _subtotal_style())
    _add(wb, _total_style())
    _add(wb, _grand_total_style())
    _add(wb, _number_style())
    _add(wb, _number_indent_style())
    _add(wb, _subtotal_number_style())
    _add(wb, _total_number_style())
    _add(wb, _grand_total_number_style())
    _add(wb, _col_header_style())
    _add(wb, _reconciled_style())
    _add(wb, _difference_style())


def _add(wb, style: NamedStyle) -> None:
    if style.name not in wb.named_styles:
        wb.add_named_style(style)


# ── Style definitions ─────────────────────────────────────────────────────────

def _title_style() -> NamedStyle:
    s = NamedStyle(name="title")
    s.font = Font(name="Calibri", bold=True, size=14, color=COLOR_DARK_BLUE)
    s.alignment = Alignment(horizontal="left", vertical="center")
    return s


def _report_header_style() -> NamedStyle:
    s = NamedStyle(name="report_header")
    s.font = Font(name="Calibri", bold=False, size=11, color="595959")
    s.alignment = Alignment(horizontal="left", vertical="center")
    return s


def _section_header_style() -> NamedStyle:
    s = NamedStyle(name="section_header")
    s.font = Font(name="Calibri", bold=True, size=11, color=COLOR_WHITE)
    s.fill = PatternFill(
        start_color=COLOR_MID_BLUE, end_color=COLOR_MID_BLUE, fill_type="solid"
    )
    s.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    return s


def _body_style() -> NamedStyle:
    s = NamedStyle(name="body")
    s.font = Font(name="Calibri", size=10)
    s.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    return s


def _body_indent_style() -> NamedStyle:
    s = NamedStyle(name="body_indent")
    s.font = Font(name="Calibri", size=10)
    s.alignment = Alignment(horizontal="left", vertical="center", indent=4)
    return s


def _subtotal_style() -> NamedStyle:
    s = NamedStyle(name="subtotal_label")
    s.font = Font(name="Calibri", bold=True, size=10)
    s.fill = PatternFill(
        start_color=COLOR_LIGHT_BLUE, end_color=COLOR_LIGHT_BLUE, fill_type="solid"
    )
    s.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    s.border = Border(top=THIN, bottom=THIN)
    return s


def _total_style() -> NamedStyle:
    s = NamedStyle(name="total_label")
    s.font = Font(name="Calibri", bold=True, size=11)
    s.fill = PatternFill(
        start_color=COLOR_LIGHT_BLUE, end_color=COLOR_LIGHT_BLUE, fill_type="solid"
    )
    s.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    s.border = Border(top=DOUBLE, bottom=DOUBLE)
    return s


def _grand_total_style() -> NamedStyle:
    s = NamedStyle(name="grand_total_label")
    s.font = Font(name="Calibri", bold=True, size=12)
    s.fill = PatternFill(
        start_color=COLOR_DARK_BLUE, end_color=COLOR_DARK_BLUE, fill_type="solid"
    )
    s.font = Font(name="Calibri", bold=True, size=12, color=COLOR_WHITE)
    s.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    s.border = Border(top=DOUBLE, bottom=DOUBLE)
    return s


def _number_style() -> NamedStyle:
    s = NamedStyle(name="number")
    s.font = Font(name="Calibri", size=10)
    s.number_format = FMT_ACCOUNTING
    s.alignment = Alignment(horizontal="right", vertical="center")
    return s


def _number_indent_style() -> NamedStyle:
    s = NamedStyle(name="number_indent")
    s.font = Font(name="Calibri", size=10)
    s.number_format = FMT_ACCOUNTING
    s.alignment = Alignment(horizontal="right", vertical="center")
    return s


def _subtotal_number_style() -> NamedStyle:
    s = NamedStyle(name="subtotal_number")
    s.font = Font(name="Calibri", bold=True, size=10)
    s.number_format = FMT_ACCOUNTING
    s.fill = PatternFill(
        start_color=COLOR_LIGHT_BLUE, end_color=COLOR_LIGHT_BLUE, fill_type="solid"
    )
    s.alignment = Alignment(horizontal="right", vertical="center")
    s.border = Border(top=THIN, bottom=THIN)
    return s


def _total_number_style() -> NamedStyle:
    s = NamedStyle(name="total_number")
    s.font = Font(name="Calibri", bold=True, size=11)
    s.number_format = FMT_ACCOUNTING
    s.fill = PatternFill(
        start_color=COLOR_LIGHT_BLUE, end_color=COLOR_LIGHT_BLUE, fill_type="solid"
    )
    s.alignment = Alignment(horizontal="right", vertical="center")
    s.border = Border(top=DOUBLE, bottom=DOUBLE)
    return s


def _grand_total_number_style() -> NamedStyle:
    s = NamedStyle(name="grand_total_number")
    s.font = Font(name="Calibri", bold=True, size=12, color=COLOR_WHITE)
    s.number_format = FMT_ACCOUNTING
    s.fill = PatternFill(
        start_color=COLOR_DARK_BLUE, end_color=COLOR_DARK_BLUE, fill_type="solid"
    )
    s.alignment = Alignment(horizontal="right", vertical="center")
    s.border = Border(top=DOUBLE, bottom=DOUBLE)
    return s


def _col_header_style() -> NamedStyle:
    s = NamedStyle(name="col_header")
    s.font = Font(name="Calibri", bold=True, size=10, color=COLOR_WHITE)
    s.fill = PatternFill(
        start_color=COLOR_DARK_BLUE, end_color=COLOR_DARK_BLUE, fill_type="solid"
    )
    s.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    s.border = Border(bottom=THIN)
    return s


def _reconciled_style() -> NamedStyle:
    s = NamedStyle(name="reconciled")
    s.font = Font(name="Calibri", bold=True, size=10, color=COLOR_GREEN)
    s.number_format = FMT_ACCOUNTING
    s.alignment = Alignment(horizontal="right", vertical="center")
    return s


def _difference_style() -> NamedStyle:
    s = NamedStyle(name="difference")
    s.font = Font(name="Calibri", bold=True, size=10, color=COLOR_RED)
    s.number_format = FMT_ACCOUNTING
    s.alignment = Alignment(horizontal="right", vertical="center")
    return s
