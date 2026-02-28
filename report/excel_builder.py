"""
Excel report builder — creates the multi-tab cash flow workbook.

Tabs produced:
  1. Cash Flow    — Formatted indirect-method cash flow statement
  2. WC Detail    — Working capital account-level detail
  3. Cash Txns    — GL transaction detail for cash/bank accounts
  4. Reconciliation — CFS vs GL ending cash balance check
  5. Audit Trail  — Raw data used (optional, controlled by config)
"""

from __future__ import annotations

import os
from datetime import datetime
from typing import TYPE_CHECKING

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from report.formatters import (
    COLOR_DARK_BLUE,
    COLOR_LIGHT_BLUE,
    COLOR_MID_BLUE,
    COLOR_RED,
    COLOR_GREEN,
    COLOR_WHITE,
    FMT_ACCOUNTING,
    FMT_DATE,
    register_styles,
)

if TYPE_CHECKING:
    from data.models import CashFlowStatement, ExtractedData, Transaction
    from cashflow.reconciler import ReconciliationResult


class ExcelReportBuilder:
    def __init__(self, config: dict) -> None:
        self.config = config
        self._output_cfg = config.get("output", {})

    def build(
        self,
        cfs: "CashFlowStatement",
        data: "ExtractedData",
        recon: "ReconciliationResult",
        output_path: str,
    ) -> str:
        """
        Build the Excel workbook and save it to output_path.

        Returns:
            The absolute path to the saved file.
        """
        wb = Workbook()
        register_styles(wb)

        # Remove default sheet
        default_sheet = wb.active
        wb.remove(default_sheet)

        self._build_cash_flow_tab(wb, cfs, recon)
        self._build_wc_detail_tab(wb, cfs, data)
        self._build_cash_txns_tab(wb, data)
        self._build_reconciliation_tab(wb, cfs, recon)

        if self._output_cfg.get("include_audit_trail", True):
            self._build_audit_trail_tab(wb, cfs, data)

        # Ensure output directory exists
        os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
        wb.save(output_path)
        return os.path.abspath(output_path)

    # ── Tab 1: Cash Flow Statement ────────────────────────────────────────────

    def _build_cash_flow_tab(self, wb, cfs, recon) -> None:
        ws = wb.create_sheet("Cash Flow")
        ws.sheet_view.showGridLines = False

        # Column widths
        ws.column_dimensions["A"].width = 50
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 5   # spacer

        row = 1

        # Title block
        company = cfs.company_name or self._output_cfg.get("company_name", "")
        _write(ws, row, 1, company or "Cash Flow Statement", style="title", col_span=2)
        row += 1
        _write(ws, row, 1, "Statement of Cash Flows (Indirect Method)", style="report_header")
        row += 1
        _write(ws, row, 1, f"For the Month Ended {_period_end_label(cfs.period)}", style="report_header")
        row += 1
        _write(ws, row, 1, f"Generated: {datetime.now().strftime('%B %d, %Y')}", style="report_header")
        row += 2   # blank row

        # Freeze panes below title block
        ws.freeze_panes = f"A{row}"

        # ── Operating Activities ──────────────────────────────────────────────
        _write(ws, row, 1, "OPERATING ACTIVITIES", style="section_header")
        _fill_row(ws, row, COLOR_MID_BLUE, cols=2)
        row += 1

        _write(ws, row, 1, "Net Income", style="body")
        _write(ws, row, 2, cfs.net_income, style="number")
        row += 1

        if cfs.noncash_items:
            _write(ws, row, 1, "Adjustments to reconcile net income to cash:", style="body")
            row += 1
            for item in cfs.noncash_items:
                _write(ws, row, 1, f"  {item.label}", style="body_indent")
                _write(ws, row, 2, item.amount, style="number_indent")
                row += 1

        if cfs.wc_changes:
            _write(ws, row, 1, "Changes in working capital:", style="body")
            row += 1
            for item in cfs.wc_changes:
                _write(ws, row, 1, f"  {_wc_label(item)}", style="body_indent")
                _write(ws, row, 2, item.amount, style="number_indent")
                row += 1

        for item in cfs.custom_operating:
            _write(ws, row, 1, f"  {item.label}", style="body_indent")
            _write(ws, row, 2, item.amount, style="number_indent")
            row += 1

        _write(ws, row, 1, "Net Cash from Operating Activities", style="total_label")
        _write(ws, row, 2, cfs.operating_total, style="total_number")
        _fill_row(ws, row, COLOR_LIGHT_BLUE, cols=2)
        row += 2

        # ── Investing Activities ──────────────────────────────────────────────
        _write(ws, row, 1, "INVESTING ACTIVITIES", style="section_header")
        _fill_row(ws, row, COLOR_MID_BLUE, cols=2)
        row += 1

        if cfs.investing_items:
            for item in cfs.investing_items:
                _write(ws, row, 1, f"  {item.label}", style="body_indent")
                _write(ws, row, 2, item.amount, style="number_indent")
                row += 1
        else:
            _write(ws, row, 1, "  No investing activity this period", style="body_indent")
            row += 1

        _write(ws, row, 1, "Net Cash from Investing Activities", style="total_label")
        _write(ws, row, 2, cfs.investing_total, style="total_number")
        _fill_row(ws, row, COLOR_LIGHT_BLUE, cols=2)
        row += 2

        # ── Financing Activities ──────────────────────────────────────────────
        _write(ws, row, 1, "FINANCING ACTIVITIES", style="section_header")
        _fill_row(ws, row, COLOR_MID_BLUE, cols=2)
        row += 1

        if cfs.financing_items:
            for item in cfs.financing_items:
                _write(ws, row, 1, f"  {item.label}", style="body_indent")
                _write(ws, row, 2, item.amount, style="number_indent")
                row += 1
        else:
            _write(ws, row, 1, "  No financing activity this period", style="body_indent")
            row += 1

        _write(ws, row, 1, "Net Cash from Financing Activities", style="total_label")
        _write(ws, row, 2, cfs.financing_total, style="total_number")
        _fill_row(ws, row, COLOR_LIGHT_BLUE, cols=2)
        row += 2

        # ── Net Change ────────────────────────────────────────────────────────
        _write(ws, row, 1, "NET INCREASE (DECREASE) IN CASH", style="grand_total_label")
        _write(ws, row, 2, cfs.net_change_in_cash, style="grand_total_number")
        _fill_row(ws, row, COLOR_DARK_BLUE, cols=2)
        row += 1

        _write(ws, row, 1, "Cash and Cash Equivalents — Beginning of Period", style="body")
        _write(ws, row, 2, cfs.beginning_cash, style="number")
        row += 1

        _write(ws, row, 1, "Cash and Cash Equivalents — End of Period", style="grand_total_label")
        _write(ws, row, 2, cfs.ending_cash_statement, style="grand_total_number")
        _fill_row(ws, row, COLOR_DARK_BLUE, cols=2)
        row += 2

        # ── Reconciliation check ──────────────────────────────────────────────
        _write(ws, row, 1, "GL Ending Cash Balance (per NetSuite)", style="body")
        _write(ws, row, 2, cfs.ending_cash_gl, style="number")
        row += 1

        diff_style = "reconciled" if recon.is_reconciled else "difference"
        diff_label = "Reconciliation Difference" if not recon.is_reconciled else "Reconciled ✓"
        _write(ws, row, 1, diff_label, style=diff_style)
        _write(ws, row, 2, recon.difference, style=diff_style)
        row += 1

    # ── Tab 2: Working Capital Detail ─────────────────────────────────────────

    def _build_wc_detail_tab(self, wb, cfs, data) -> None:
        ws = wb.create_sheet("WC Detail")
        ws.sheet_view.showGridLines = False

        headers = [
            "Account Name", "Account Type",
            "Prior Period Balance", "Current Period Balance",
            "Balance Change", "Cash Flow Impact",
        ]
        col_widths = [40, 18, 22, 22, 18, 18]
        for i, (h, w) in enumerate(zip(headers, col_widths), 1):
            ws.column_dimensions[get_column_letter(i)].width = w
            _write(ws, 1, i, h, style="col_header")

        ws.row_dimensions[1].height = 30
        ws.freeze_panes = "A2"

        row = 2
        for item in cfs.wc_changes:
            # Find matching account data
            prior_bal = 0.0
            curr_bal = 0.0
            if item.account_id:
                for a in data.bs_prior:
                    if a.account_id == item.account_id:
                        prior_bal = a.natural_balance
                        break
                for a in data.bs_current:
                    if a.account_id == item.account_id:
                        curr_bal = a.natural_balance
                        break

            change = curr_bal - prior_bal
            _write(ws, row, 1, item.label)
            _write(ws, row, 2, item.account_type or "")
            _write_num(ws, row, 3, prior_bal)
            _write_num(ws, row, 4, curr_bal)
            _write_num(ws, row, 5, change)
            _write_num(ws, row, 6, item.amount)
            row += 1

        # Totals row
        row += 1
        _write(ws, row, 1, "Total Working Capital Cash Impact", style="subtotal_label")
        _write(ws, row, 6, cfs.wc_total, style="subtotal_number")

    # ── Tab 3: Cash Transactions ──────────────────────────────────────────────

    def _build_cash_txns_tab(self, wb, data) -> None:
        ws = wb.create_sheet("Cash Txns")
        ws.sheet_view.showGridLines = False

        headers = ["Date", "Type", "Reference", "Entity", "Account", "Memo", "Debit", "Credit", "Net Amount"]
        col_widths = [13, 16, 14, 28, 28, 36, 14, 14, 14]
        for i, (h, w) in enumerate(zip(headers, col_widths), 1):
            ws.column_dimensions[get_column_letter(i)].width = w
            _write(ws, 1, i, h, style="col_header")

        ws.row_dimensions[1].height = 25
        ws.freeze_panes = "A2"

        # Enable autofilter
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

        row = 2
        running = 0.0
        for txn in sorted(data.cash_transactions, key=lambda t: t.date):
            net = txn.debit - txn.credit
            running += net
            ws.cell(row, 1, txn.date)
            ws.cell(row, 2, txn.transaction_type)
            ws.cell(row, 3, txn.reference)
            ws.cell(row, 4, txn.entity)
            ws.cell(row, 5, txn.account_name)
            ws.cell(row, 6, txn.memo)
            _write_num(ws, row, 7, txn.debit if txn.debit else None)
            _write_num(ws, row, 8, txn.credit if txn.credit else None)
            _write_num(ws, row, 9, net)
            row += 1

        # Totals
        row += 1
        _write(ws, row, 6, "Net Cash Change", style="subtotal_label")
        _write_num(ws, row, 9, running, style="subtotal_number")

    # ── Tab 4: Reconciliation ─────────────────────────────────────────────────

    def _build_reconciliation_tab(self, wb, cfs, recon) -> None:
        ws = wb.create_sheet("Reconciliation")
        ws.sheet_view.showGridLines = False
        ws.column_dimensions["A"].width = 45
        ws.column_dimensions["B"].width = 20

        row = 1
        _write(ws, row, 1, "Cash Balance Reconciliation", style="title")
        row += 1
        _write(ws, row, 1, f"Period: {cfs.period.name}", style="report_header")
        row += 2

        _write(ws, row, 1, "PER CASH FLOW STATEMENT", style="section_header")
        _fill_row(ws, row, COLOR_MID_BLUE, cols=2)
        row += 1

        _write(ws, row, 1, "Beginning Cash Balance", style="body")
        _write(ws, row, 2, cfs.beginning_cash, style="number")
        row += 1

        _write(ws, row, 1, "  + Net Cash from Operating Activities", style="body_indent")
        _write(ws, row, 2, cfs.operating_total, style="number_indent")
        row += 1

        _write(ws, row, 1, "  + Net Cash from Investing Activities", style="body_indent")
        _write(ws, row, 2, cfs.investing_total, style="number_indent")
        row += 1

        _write(ws, row, 1, "  + Net Cash from Financing Activities", style="body_indent")
        _write(ws, row, 2, cfs.financing_total, style="number_indent")
        row += 1

        _write(ws, row, 1, "Ending Cash (per Statement)", style="total_label")
        _write(ws, row, 2, cfs.ending_cash_statement, style="total_number")
        _fill_row(ws, row, COLOR_LIGHT_BLUE, cols=2)
        row += 2

        _write(ws, row, 1, "PER GENERAL LEDGER", style="section_header")
        _fill_row(ws, row, COLOR_MID_BLUE, cols=2)
        row += 1

        _write(ws, row, 1, "Ending Cash Balance (sum of bank accounts in GL)", style="body")
        _write(ws, row, 2, cfs.ending_cash_gl, style="number")
        row += 2

        diff_style = "reconciled" if recon.is_reconciled else "difference"
        _write(ws, row, 1, "DIFFERENCE (GL − Statement)", style=diff_style)
        _write(ws, row, 2, recon.difference, style=diff_style)
        row += 1

        status = "RECONCILED" if recon.is_reconciled else "DIFFERENCE — SEE NOTE BELOW"
        _write(ws, row, 1, f"Status: {status}", style=diff_style)
        row += 2

        if not recon.is_reconciled:
            note = (
                "A non-zero difference indicates one or more transactions are not "
                "captured in the cash flow statement sections. Common causes:\n"
                "  • An account is not mapped in config/settings.yaml\n"
                "  • A one-off transaction requires a custom_adjustment entry\n"
                "  • Intercompany or reclassification entries without cash impact"
            )
            _write(ws, row, 1, note, style="body")
            ws.row_dimensions[row].height = 75
            ws.cell(row, 1).alignment = Alignment(wrap_text=True, vertical="top")

    # ── Tab 5: Audit Trail ────────────────────────────────────────────────────

    def _build_audit_trail_tab(self, wb, cfs, data) -> None:
        ws = wb.create_sheet("Audit Trail")
        ws.sheet_view.showGridLines = False

        row = 1
        _write(ws, row, 1, "Audit Trail — Raw Data Used to Compute Cash Flow", style="title")
        row += 2

        # Run metadata
        _write(ws, row, 1, "Run Information", style="section_header")
        _fill_row(ws, row, COLOR_MID_BLUE, cols=4)
        row += 1
        metadata = [
            ("Period", cfs.period.name),
            ("Period Start", cfs.period.start_date),
            ("Period End", cfs.period.end_date),
            ("Prior Period", data.prior_period.name),
            ("Generated At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ]
        for label, value in metadata:
            _write(ws, row, 1, label, style="body")
            _write(ws, row, 2, str(value))
            row += 1
        row += 1

        # P&L accounts
        row = _write_account_table(
            ws, row, "P&L Account Activity (Period)", data.pl_accounts
        )
        row += 1

        # Balance sheet — current
        row = _write_account_table(
            ws, row, "Balance Sheet — Current Period End", data.bs_current
        )
        row += 1

        # Balance sheet — prior
        row = _write_account_table(
            ws, row, "Balance Sheet — Prior Period End", data.bs_prior
        )

        # Column widths
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 40
        ws.column_dimensions["D"].width = 16
        ws.column_dimensions["E"].width = 16
        ws.column_dimensions["F"].width = 16


# ── Worksheet helpers ─────────────────────────────────────────────────────────

def _write(ws, row: int, col: int, value, style: str | None = None, col_span: int = 1) -> None:
    cell = ws.cell(row, col, value)
    if style:
        cell.style = style
    if col_span > 1:
        ws.merge_cells(
            start_row=row, start_column=col,
            end_row=row, end_column=col + col_span - 1,
        )


def _write_num(ws, row: int, col: int, value, style: str = "number") -> None:
    if value is None:
        return
    cell = ws.cell(row, col, value)
    cell.style = style
    cell.number_format = FMT_ACCOUNTING


def _fill_row(ws, row: int, color: str, cols: int = 10) -> None:
    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    for c in range(1, cols + 1):
        ws.cell(row, c).fill = fill


def _write_account_table(ws, start_row: int, title: str, accounts) -> int:
    from data.models import AccountBalance
    row = start_row
    _write(ws, row, 1, title, style="section_header")
    _fill_row(ws, row, COLOR_MID_BLUE, cols=6)
    row += 1

    headers = ["Acct ID", "Acct Number", "Account Name", "Type", "Debits", "Credits"]
    for c, h in enumerate(headers, 1):
        _write(ws, row, c, h, style="col_header")
    row += 1

    for a in accounts:
        ws.cell(row, 1, a.account_id)
        ws.cell(row, 2, a.account_number)
        ws.cell(row, 3, a.account_name)
        ws.cell(row, 4, a.account_type)
        _write_num(ws, row, 5, a.total_debits)
        _write_num(ws, row, 6, a.total_credits)
        row += 1

    return row


def _period_end_label(period) -> str:
    """Format period end date as 'January 31, 2026'."""
    from datetime import date
    try:
        d = date.fromisoformat(period.end_date)
        return d.strftime("%B %d, %Y")
    except Exception:
        return period.name


def _wc_label(item) -> str:
    """
    Apply the standard accounting label convention for working capital items:
      - Assets: "(Increase)/Decrease in {Name}"
      - Liabilities: "Increase/(Decrease) in {Name}"
    """
    from data.models import ASSET_WC_TYPES
    if item.account_type in ASSET_WC_TYPES:
        return f"(Increase)/Decrease in {item.label}"
    return f"Increase/(Decrease) in {item.label}"
